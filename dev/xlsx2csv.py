import os
import csv
from openpyxl import load_workbook

def convert_xlsx_sheets_to_csv(directory_path, output_dir=None):
    """
    Scans a directory for .xlsx files and converts each sheet 
    in those files into separate CSV files.

    :param directory_path: The path to the directory to scan.
    :param output_dir: Optional. The directory where CSV files should be saved.
                       If None, they are saved in the input directory.
    """
    
    # 1. Set the output directory
    if output_dir is None:
        output_dir = directory_path
    elif not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created output directory: {output_dir}")

    print(f"--- Starting Scan in: {directory_path} ---")

    # 2. Iterate through all items in the directory
    for item in os.listdir(directory_path):
        # Construct the full path
        file_path = os.path.join(directory_path, item)
        
        # 3. Check if the item is an XLSX file
        if item.endswith('.xlsx') and os.path.isfile(file_path):
            print(f"\nProcessing file: **{item}**")
            
            try:
                # Load the workbook
                workbook = load_workbook(file_path)
                
                # Get the base filename without the extension
                base_filename = os.path.splitext(item)[0]
                
                # 4. Iterate through all sheets in the workbook
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    
                    # 5. Construct the output CSV filename
                    # Format: {base_filename}-{sheet_name}.csv
                    csv_filename = f"{base_filename}-{sheet_name}.csv"
                    csv_file_path = os.path.join(output_dir, csv_filename)
                    
                    print(f"   -> Converting sheet '{sheet_name}' to **{csv_filename}**")

                    # 6. Write the sheet data to a CSV file
                    with open(csv_file_path, 'w', newline='', encoding='utf-8') as csvfile:
                        csv_writer = csv.writer(csvfile)
                        
                        for row in worksheet.iter_rows():
                            # Extract values from cells in the row
                            row_values = [cell.value for cell in row]
                            csv_writer.writerow(row_values)
                            
            except Exception as e:
                print(f"   **ERROR** processing {item}: {e}")

    print("\n--- Conversion Complete ---")


# --- Configuration ---
# **CHANGE THIS PATH** to the directory containing your .xlsx files
target_directory = './data' 

# Optional: Specify a different directory for the output CSVs
# Set to None to save them in the target_directory
csv_output_directory = None 
# ---------------------


# Ensure the target directory exists before running
if os.path.isdir(target_directory):
    convert_xlsx_sheets_to_csv(target_directory, csv_output_directory)
else:
    print(f"Error: The directory '{target_directory}' does not exist.")
